"""
Generates sample data files for TLF Output Generator:
  1. sample_shell_annotated.png  — annotated oncology efficacy mock shell
  2. sample_adam_specs.xlsx      — ADRS AdaM specs (Variable Level + Value Level)
  3. sample_adrs.csv             — 90-subject ADRS dataset (BOR + ORR + DCR)
Run once:  python data/generate_samples.py
"""

import os
import random
import csv
from pathlib import Path
from datetime import datetime, timedelta

OUT_DIR = Path(__file__).parent

# ─────────────────────────────────────────────────────────────────
# 1. Annotated PNG mock shell
# ─────────────────────────────────────────────────────────────────
def make_annotated_shell():
    from PIL import Image, ImageDraw, ImageFont

    W, H = 1100, 780
    BG          = (255, 255, 255)
    BLACK       = (0,   0,   0)
    HEADER_BG   = (30,  60, 110)
    HEADER_FG   = (255, 255, 255)
    ALT_ROW     = (240, 245, 255)
    BORDER      = (180, 190, 210)
    ANN_BOX     = (255, 80,  40)   # annotation callout colour
    ANN_TEXT    = (200, 30,  10)
    ANN_LINE    = (220, 60,  20)
    GRID_LINE   = (210, 215, 225)

    img  = Image.new("RGB", (W, H), BG)
    draw = ImageDraw.Draw(img)

    # ── try to load a decent system font; fall back gracefully ──
    def font(size, bold=False):
        candidates = [
            "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/Arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
        ]
        for path in candidates:
            if os.path.exists(path):
                return ImageFont.truetype(path, size)
        return ImageFont.load_default()

    f_title   = font(13, bold=True)
    f_sub     = font(11)
    f_hdr     = font(11, bold=True)
    f_cell    = font(10)
    f_ann     = font(10, bold=True)
    f_ann_sm  = font(9)

    # ── table geometry ──
    L, T = 60, 120          # left / top of table body
    COL_W = [220, 150, 150, 150, 150]   # stub + 4 data cols
    ROW_H = 28
    COLS  = len(COL_W)

    col_x = [L]
    for w in COL_W[:-1]:
        col_x.append(col_x[-1] + w)
    TABLE_W = sum(COL_W)

    # ── title block ──
    draw.text((L, 18), "TLF Output Generator — Sample Oncology Efficacy Shell", font=f_title, fill=BLACK)
    draw.text((L, 38), "Table 14.3.1: Summary of Tumor Response by Best Overall Response", font=f_title, fill=BLACK)
    draw.text((L, 56), "Full Analysis Set (FAS) | Data Source: ADRS", font=f_sub, fill=(80, 80, 80))
    draw.text((L, 72), "Parameter: Best Overall Response (PARAMCD = 'BOR')", font=f_sub, fill=(80, 80, 80))
    draw.line([(L, 94), (L + TABLE_W, 94)], fill=BORDER, width=1)

    # ── header row ──
    headers = ["Response Category", "Placebo\n(N=xx)", "Drug A 50mg\n(N=xx)", "Drug A 100mg\n(N=xx)", "Total\n(N=xx)"]
    draw.rectangle([L, T, L + TABLE_W, T + ROW_H * 2], fill=HEADER_BG)
    for i, (hdr, x) in enumerate(zip(headers, col_x)):
        lines = hdr.split("\n")
        for j, line in enumerate(lines):
            draw.text((x + 6, T + 4 + j * 14), line, font=f_hdr, fill=HEADER_FG)

    # ── data rows ──
    rows = [
        ("Best Overall Response",             "",         "",         "",         ""),
        ("  Complete Response (CR)",           "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Partial Response (PR)",            "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Stable Disease (SD)",              "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Progressive Disease (PD)",         "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Not Evaluable / Missing",          "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("Overall Response Rate (CR+PR)",      "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  95% CI (Clopper-Pearson)",         "(x.x, x.x)","(x.x, x.x)","(x.x, x.x)","(x.x, x.x)"),
        ("Disease Control Rate (CR+PR+SD)",    "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  95% CI (Clopper-Pearson)",         "(x.x, x.x)","(x.x, x.x)","(x.x, x.x)","(x.x, x.x)"),
        ("Duration of Response (months)",      "",         "",         "",         ""),
        ("  Median (95% CI)",                  "x.x (x.x, x.x)","x.x (x.x, x.x)","x.x (x.x, x.x)","x.x (x.x, x.x)"),
    ]

    body_top = T + ROW_H * 2
    for r_idx, row in enumerate(rows):
        y0 = body_top + r_idx * ROW_H
        y1 = y0 + ROW_H
        bg = ALT_ROW if r_idx % 2 == 1 else BG
        draw.rectangle([L, y0, L + TABLE_W, y1], fill=bg)
        for c_idx, (cell, x) in enumerate(zip(row, col_x)):
            bold = (c_idx == 0 and not row[0].startswith("  "))
            draw.text((x + 6, y0 + 6), cell, font=f_hdr if bold else f_cell, fill=BLACK)

    # ── grid lines ──
    total_rows = len(rows)
    table_bottom = body_top + total_rows * ROW_H
    # horizontal
    for r in range(total_rows + 1):
        y = body_top + r * ROW_H
        draw.line([(L, y), (L + TABLE_W, y)], fill=GRID_LINE, width=1)
    # vertical
    for x in col_x:
        draw.line([(x, T), (x, table_bottom)], fill=GRID_LINE, width=1)
    draw.line([(L + TABLE_W, T), (L + TABLE_W, table_bottom)], fill=GRID_LINE, width=1)
    # outer border
    draw.rectangle([L, T, L + TABLE_W, table_bottom], outline=BORDER, width=2)

    # ── annotations ──
    def callout(ax, ay, bx, by, label, sub=""):
        """Draw annotation arrow from (ax,ay) to (bx,by) with a label box."""
        draw.line([(ax, ay), (bx, by)], fill=ANN_LINE, width=2)
        draw.ellipse([(bx-4, by-4), (bx+4, by+4)], fill=ANN_LINE)
        # box
        bw = max(len(label), len(sub)) * 7 + 10
        bh = 30 if sub else 18
        draw.rectangle([(ax-4, ay-bh), (ax+bw, ay+4)], fill=(255, 245, 240), outline=ANN_BOX, width=1)
        draw.text((ax, ay-bh+3), label, font=f_ann, fill=ANN_TEXT)
        if sub:
            draw.text((ax, ay-bh+16), sub, font=f_ann_sm, fill=ANN_TEXT)

    # Annotation 1: Dataset
    callout(ax=720, ay=T-12, bx=col_x[0]+110, by=T+12,
            label="Dataset: ADRS",
            sub="adam_specs → dataset_source")

    # Annotation 2: PARAMCD filter
    callout(ax=750, ay=T+62, bx=col_x[0]+60, by=body_top+4,
            label="Filter: PARAMCD = 'BOR'",
            sub="(Best Overall Response record)")

    # Annotation 3: AVALC variable
    callout(ax=730, ay=body_top+90, bx=col_x[0]+50, by=body_top + 1*ROW_H + 14,
            label="analysis_var: AVALC",
            sub="Response category (CR/PR/SD/PD)")

    # Annotation 4: ORR derived row
    callout(ax=730, ay=body_top+200, bx=col_x[0]+80, by=body_top + 6*ROW_H + 14,
            label="Derived: ORR = AVALC in (CR, PR)",
            sub="Condition from adam_specs codelist")

    # Annotation 5: Treatment column
    callout(ax=800, ay=H-60, bx=col_x[2]+75, by=T+10,
            label="Treatment var: TRTP",
            sub="(planned treatment variable)")

    # ── footnote ──
    draw.line([(L, table_bottom + 10), (L + TABLE_W, table_bottom + 10)], fill=BORDER, width=1)
    draw.text((L, table_bottom + 16), "Note: Percentages based on number of subjects in each treatment group (N).", font=f_ann_sm, fill=(80,80,80))
    draw.text((L, table_bottom + 30), "CI = Confidence Interval; CR = Complete Response; PR = Partial Response; SD = Stable Disease; PD = Progressive Disease.", font=f_ann_sm, fill=(80,80,80))

    # ── legend box ──
    lx, ly = L, H - 55
    draw.rectangle([(lx, ly), (lx+500, ly+45)], fill=(255,250,240), outline=ANN_BOX, width=1)
    draw.text((lx+6, ly+4),  "ANNOTATION LEGEND", font=f_ann, fill=ANN_TEXT)
    draw.text((lx+6, ly+18), "Red callouts = metadata used to generate R code (dataset, variables, filters, conditions)", font=f_ann_sm, fill=ANN_TEXT)
    draw.text((lx+6, ly+32), "Upload this shell in Step 1 · Upload AdaM specs in Step 2 · Generate R code in Step 4", font=f_ann_sm, fill=ANN_TEXT)

    out = OUT_DIR / "sample_shell_annotated.png"
    img.save(str(out), dpi=(150, 150))
    print(f"Saved: {out}")


# ─────────────────────────────────────────────────────────────────
# 2. Sample AdaM specs Excel — Variable Level + Value Level
# ─────────────────────────────────────────────────────────────────
def make_adam_specs():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── helper styles ──
    HDR_FILL   = PatternFill("solid", fgColor="1E3C6E")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    CELL_FONT  = Font(size=9)
    ALT_FILL   = PatternFill("solid", fgColor="F2F6FC")
    thin = Side(style="thin", color="B0B8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def hdr_row(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = wrap; c.border = border

    def data_row(ws, row, values, alt=False):
        fill = ALT_FILL if alt else PatternFill()
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = CELL_FONT; c.alignment = wrap; c.border = border
            if alt: c.fill = fill

    # ══════════════════════════════════════════════════════════════
    # Sheet 1: Variable Level
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Variable_Level"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value = "ADRS — Variable Level Metadata"
    t.font = Font(bold=True, size=12, color="1E3C6E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    ws1.merge_cells("A2:H2")
    t2 = ws1["A2"]
    t2.value = "Dataset: ADRS (Oncology Response) | Standard: CDISC ADaM v1.0 | Population: FASFL = 'Y'"
    t2.font = Font(italic=True, size=10, color="555555")
    t2.alignment = Alignment(horizontal="center")

    cols = ["Variable", "Label", "Type", "Length", "Codelist / Values", "ADaM Core", "Key Flag", "Derivation / Notes"]
    hdr_row(ws1, 3, cols)

    variables = [
        ("STUDYID",  "Study Identifier",              "Char", 20,  "",                          "Req",    "",    "From SDTM DM.STUDYID"),
        ("USUBJID",  "Unique Subject Identifier",     "Char", 40,  "",                          "Req",    "",    "From SDTM DM.USUBJID — primary merge key to ADSL"),
        ("SUBJID",   "Subject Identifier",            "Char", 20,  "",                          "Req",    "",    "From SDTM DM.SUBJID"),
        ("SITEID",   "Study Site Identifier",         "Char", 20,  "",                          "Perm",   "",    "From SDTM DM.SITEID"),
        ("TRTP",     "Planned Treatment",             "Char", 200, "Placebo / Drug A 50mg / Drug A 100mg", "Req", "", "Mapped from ADSL.TRT01P — use for column grouping in tables"),
        ("TRTA",     "Actual Treatment",              "Char", 200, "Placebo / Drug A 50mg / Drug A 100mg", "Req", "", "Mapped from ADSL.TRT01A — use for safety tables"),
        ("PARAMCD",  "Parameter Code",                "Char", 8,   "BOR / ORR / DCR",           "Req",    "KEY", "Identifies the analysis parameter — filter on this variable"),
        ("PARAM",    "Parameter Description",         "Char", 200, "",                          "Req",    "",    "Long description of PARAMCD"),
        ("AVAL",     "Analysis Value (numeric)",      "Num",  8,   "1 = Yes/Responder, 0 = No/Non-resp", "Req", "",  "Numeric version of response flag — see Value Level for derivation"),
        ("AVALC",    "Analysis Value (character)",    "Char", 200, "CR / PR / SD / PD / NE",    "Req",    "KEY", "Primary analysis variable — see Value Level for derivation per PARAMCD"),
        ("ANL01FL",  "Analysis Record Flag",          "Char", 1,   "Y / N",                     "Req",    "KEY", "Y = include in primary analysis. Always filter: ANL01FL = 'Y'"),
        ("FASFL",    "Full Analysis Set Flag",        "Char", 1,   "Y / N",                     "Req",    "KEY", "Y = subject in FAS population. Always filter: FASFL = 'Y'"),
        ("ADT",      "Analysis Date",                 "Num",  8,   "",                          "Perm",   "",    "Date of tumor assessment (SAS date)"),
    ]

    for i, row in enumerate(variables):
        data_row(ws1, i + 4, row, alt=(i % 2 == 1))

    widths = [12, 28, 6, 8, 38, 10, 8, 54]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ══════════════════════════════════════════════════════════════
    # Sheet 2: Value Level — derivation per PARAMCD
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Value_Level")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:I1")
    t = ws2["A1"]
    t.value = "ADRS — Value Level Metadata (Derivation per PARAMCD)"
    t.font = Font(bold=True, size=12, color="1E3C6E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells("A2:I2")
    t2 = ws2["A2"]
    t2.value = "Each row describes how AVALC and AVAL are derived for a given PARAMCD value"
    t2.font = Font(italic=True, size=10, color="555555")
    t2.alignment = Alignment(horizontal="center")

    hdr_row(ws2, 3, [
        "PARAMCD", "PARAM (Description)", "Source",
        "AVALC Derivation", "AVALC Possible Values",
        "AVAL Derivation",
        "Population Filter", "Analysis Flag", "Table Reference",
    ])

    value_rows = [
        (
            "BOR",
            "Best Overall Response",
            "SDTM RS / TU domains",
            "Best confirmed response per RECIST 1.1: take the best response "
            "across all post-baseline tumor assessments. CR > PR > SD > PD > NE.",
            "CR = Complete Response\n"
            "PR = Partial Response\n"
            "SD = Stable Disease\n"
            "PD = Progressive Disease\n"
            "NE = Not Evaluable",
            "AVAL = 1 if AVALC in ('CR','PR') else 0\n(binary responder flag)",
            "FASFL = 'Y'",
            "ANL01FL = 'Y'",
            "Table 14.3.1\nTumor Response",
        ),
        (
            "ORR",
            "Overall Response Rate",
            "Derived from BOR record",
            "If subject's BOR AVALC in ('CR','PR') then AVALC = 'Responder', "
            "else AVALC = 'Non-Responder'.\nOne record per subject.",
            "Responder\nNon-Responder",
            "AVAL = 1 if Responder, 0 if Non-Responder",
            "FASFL = 'Y'",
            "ANL01FL = 'Y'",
            "Table 14.3.2\nOverall Response Rate",
        ),
        (
            "DCR",
            "Disease Control Rate",
            "Derived from BOR record",
            "If subject's BOR AVALC in ('CR','PR','SD') then AVALC = 'Controlled', "
            "else AVALC = 'Not Controlled'.\nOne record per subject.",
            "Controlled\nNot Controlled",
            "AVAL = 1 if Controlled, 0 if Not Controlled",
            "FASFL = 'Y'",
            "ANL01FL = 'Y'",
            "Table 14.3.3\nDisease Control Rate",
        ),
    ]

    for i, row in enumerate(value_rows):
        data_row(ws2, i + 4, row, alt=(i % 2 == 1))
        ws2.row_dimensions[i + 4].height = 72  # taller rows for multiline text

    ws2_widths = [10, 24, 20, 44, 28, 32, 14, 14, 18]
    for col, w in enumerate(ws2_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    out = OUT_DIR / "sample_adam_specs.xlsx"
    wb.save(str(out))
    print(f"Saved: {out}")


# ─────────────────────────────────────────────────────────────────
# 3. Sample ADRS CSV — 90 subjects × 3 arms × (BOR + ORR + DCR)
# ─────────────────────────────────────────────────────────────────
def make_adrs_csv():
    random.seed(42)

    arms = ["Placebo"] * 30 + ["Drug A 50mg"] * 30 + ["Drug A 100mg"] * 30
    sites = ["001", "002", "003", "004"]
    bor_weights = {
        "Placebo":        {"CR": 0.03, "PR": 0.10, "SD": 0.30, "PD": 0.45, "NE": 0.12},
        "Drug A 50mg":    {"CR": 0.08, "PR": 0.22, "SD": 0.35, "PD": 0.25, "NE": 0.10},
        "Drug A 100mg":   {"CR": 0.15, "PR": 0.28, "SD": 0.30, "PD": 0.18, "NE": 0.09},
    }

    def pick_bor(arm):
        w = bor_weights[arm]
        return random.choices(list(w.keys()), weights=list(w.values()), k=1)[0]

    header = [
        "STUDYID", "DOMAIN", "USUBJID", "SUBJID", "SITEID",
        "TRTP", "TRTA", "PARAMCD", "PARAM", "AVAL", "AVALC",
        "ANL01FL", "FASFL", "DTYPE", "VISITNUM", "VISIT", "ADT", "ADTM",
    ]

    rows = []
    base_date = datetime(2023, 10, 1)

    for subj_idx, arm in enumerate(arms, 1):
        subjid = f"{subj_idx:04d}"
        site = random.choice(sites)
        usubjid = f"MYSTUDY01-{site}-{subjid}"
        adt = base_date + timedelta(days=random.randint(0, 60))
        adt_str = adt.strftime("%Y-%m-%d")
        adtm_str = adt.strftime("%Y-%m-%dT%H:%M:00")

        # BOR record
        bor = pick_bor(arm)
        bor_aval = 1 if bor in ("CR", "PR") else 0
        rows.append([
            "MYSTUDY01", "ADRS", usubjid, subjid, site,
            arm, arm, "BOR", "Best Overall Response",
            bor_aval, bor, "Y", "Y", "", "99", "End of Treatment",
            adt_str, adtm_str,
        ])

        # ORR record (derived from BOR)
        orr_c = "Responder" if bor in ("CR", "PR") else "Non-Responder"
        orr_n = 1 if bor in ("CR", "PR") else 0
        rows.append([
            "MYSTUDY01", "ADRS", usubjid, subjid, site,
            arm, arm, "ORR", "Overall Response Rate",
            orr_n, orr_c, "Y", "Y", "", "99", "End of Treatment",
            adt_str, adtm_str,
        ])

        # DCR record (derived from BOR)
        dcr_c = "Controlled" if bor in ("CR", "PR", "SD") else "Not Controlled"
        dcr_n = 1 if bor in ("CR", "PR", "SD") else 0
        rows.append([
            "MYSTUDY01", "ADRS", usubjid, subjid, site,
            arm, arm, "DCR", "Disease Control Rate",
            dcr_n, dcr_c, "Y", "Y", "", "99", "End of Treatment",
            adt_str, adtm_str,
        ])

    out = OUT_DIR / "sample_adrs.csv"
    with open(out, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print(f"Saved: {out}  ({len(rows)} records, {len(arms)} subjects)")


# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    make_annotated_shell()
    make_adam_specs()
    make_adrs_csv()
    print("Sample files created successfully.")
