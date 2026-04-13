"""
Generates sample data files for TLF Output Generator:
  1. sample_shell_annotated.png  — annotated AE SOC-by-PT mock shell
  2. sample_adae_spec.xlsx       — ADAE AdaM specs (Variable Level)
  3. sample_adae.csv             — 90-subject ADAE dataset
Run once:  python data/generate_samples.py
"""

import os
import random
import csv
from pathlib import Path
from datetime import datetime, timedelta

OUT_DIR = Path(__file__).parent

# ─────────────────────────────────────────────────────────────────
# 1. Annotated PNG mock shell — AE SOC by PT
# ─────────────────────────────────────────────────────────────────
def make_annotated_shell():
    from PIL import Image, ImageDraw, ImageFont

    W, H = 1100, 780
    BG          = (255, 255, 255)
    BLACK       = (0,   0,   0)
    HEADER_BG   = (30,  60, 110)
    HEADER_FG   = (255, 255, 255)
    ALT_ROW     = (240, 245, 255)
    BORDER      = (180, 190, 210)
    ANN_BOX     = (255, 80,  40)
    ANN_TEXT    = (200, 30,  10)
    ANN_LINE    = (220, 60,  20)
    GRID_LINE   = (210, 215, 225)

    img  = Image.new("RGB", (W, H), BG)
    draw = ImageDraw.Draw(img)

    def font(size, bold=False):
        candidates = [
            "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/Arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
        ]
        for path in candidates:
            if os.path.exists(path):
                return ImageFont.truetype(path, size)
        return ImageFont.load_default()

    f_title   = font(13, bold=True)
    f_sub     = font(11)
    f_hdr     = font(11, bold=True)
    f_cell    = font(10)
    f_ann     = font(10, bold=True)
    f_ann_sm  = font(9)

    # ── table geometry ──
    L, T = 60, 120
    COL_W = [280, 150, 150, 150, 150]
    ROW_H = 26
    COLS  = len(COL_W)

    col_x = [L]
    for w in COL_W[:-1]:
        col_x.append(col_x[-1] + w)
    TABLE_W = sum(COL_W)

    # ── title block ──
    draw.text((L, 18), "TLF Output Generator — Sample Adverse Events Shell", font=f_title, fill=BLACK)
    draw.text((L, 38), "Table 14.3.1: Summary of Adverse Events by System Organ Class and Preferred Term", font=f_title, fill=BLACK)
    draw.text((L, 56), "Safety Analysis Set (SAFFL = 'Y') | Data Source: ADAE", font=f_sub, fill=(80, 80, 80))
    draw.text((L, 72), "Treatment-Emergent Adverse Events (TRTEMFL = 'Y')", font=f_sub, fill=(80, 80, 80))
    draw.line([(L, 94), (L + TABLE_W, 94)], fill=BORDER, width=1)

    # ── header row ──
    headers = ["System Organ Class\n  Preferred Term", "Placebo\n(N=xx)", "Drug A 50mg\n(N=xx)", "Drug A 100mg\n(N=xx)", "Total\n(N=xx)"]
    draw.rectangle([L, T, L + TABLE_W, T + ROW_H * 2], fill=HEADER_BG)
    for i, (hdr, x) in enumerate(zip(headers, col_x)):
        lines = hdr.split("\n")
        for j, line in enumerate(lines):
            draw.text((x + 6, T + 4 + j * 14), line, font=f_hdr, fill=HEADER_FG)

    # ── data rows ──
    rows = [
        ("Any adverse event",                  "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("Gastrointestinal disorders",         "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Nausea",                           "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Diarrhea",                         "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Vomiting",                         "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("General disorders",                  "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Fatigue",                          "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Pyrexia",                          "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("Nervous system disorders",           "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Headache",                         "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Dizziness",                        "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("Skin and subcutaneous disorders",    "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
        ("  Rash",                             "x (x.x%)", "x (x.x%)", "x (x.x%)", "x (x.x%)"),
    ]

    body_top = T + ROW_H * 2
    for r_idx, row in enumerate(rows):
        y0 = body_top + r_idx * ROW_H
        y1 = y0 + ROW_H
        bg = ALT_ROW if r_idx % 2 == 1 else BG
        draw.rectangle([L, y0, L + TABLE_W, y1], fill=bg)
        for c_idx, (cell, x) in enumerate(zip(row, col_x)):
            bold = (c_idx == 0 and not row[0].startswith("  "))
            draw.text((x + 6, y0 + 5), cell, font=f_hdr if bold else f_cell, fill=BLACK)

    # ── grid lines ──
    total_rows = len(rows)
    table_bottom = body_top + total_rows * ROW_H
    for r in range(total_rows + 1):
        y = body_top + r * ROW_H
        draw.line([(L, y), (L + TABLE_W, y)], fill=GRID_LINE, width=1)
    for x in col_x:
        draw.line([(x, T), (x, table_bottom)], fill=GRID_LINE, width=1)
    draw.line([(L + TABLE_W, T), (L + TABLE_W, table_bottom)], fill=GRID_LINE, width=1)
    draw.rectangle([L, T, L + TABLE_W, table_bottom], outline=BORDER, width=2)

    # ── annotations ──
    def callout(ax, ay, bx, by, label, sub=""):
        draw.line([(ax, ay), (bx, by)], fill=ANN_LINE, width=2)
        draw.ellipse([(bx-4, by-4), (bx+4, by+4)], fill=ANN_LINE)
        bw = max(len(label), len(sub)) * 7 + 10
        bh = 30 if sub else 18
        draw.rectangle([(ax-4, ay-bh), (ax+bw, ay+4)], fill=(255, 245, 240), outline=ANN_BOX, width=1)
        draw.text((ax, ay-bh+3), label, font=f_ann, fill=ANN_TEXT)
        if sub:
            draw.text((ax, ay-bh+16), sub, font=f_ann_sm, fill=ANN_TEXT)

    # Annotation 1: Dataset
    callout(ax=720, ay=T-12, bx=col_x[0]+110, by=T+12,
            label="Dataset: ADAE",
            sub="adam_specs -> dataset_source")

    # Annotation 2: SOC variable
    callout(ax=730, ay=body_top+50, bx=col_x[0]+80, by=body_top + 1*ROW_H + 10,
            label="SOC var: AEBODSYS",
            sub="(System Organ Class grouping)")

    # Annotation 3: PT variable
    callout(ax=730, ay=body_top+130, bx=col_x[0]+50, by=body_top + 2*ROW_H + 10,
            label="PT var: AEDECOD",
            sub="(Preferred Term — indented rows)")

    # Annotation 4: Count distinct subjects
    callout(ax=740, ay=body_top+210, bx=col_x[1]+50, by=body_top + 2*ROW_H + 10,
            label="Count: n (%) distinct USUBJID",
            sub="distinct_by = USUBJID")

    # Annotation 5: Treatment column
    callout(ax=800, ay=H-60, bx=col_x[2]+75, by=T+10,
            label="Treatment var: TRTA",
            sub="(actual treatment variable)")

    # ── footnote ──
    draw.line([(L, table_bottom + 10), (L + TABLE_W, table_bottom + 10)], fill=BORDER, width=1)
    draw.text((L, table_bottom + 16), "Note: Percentages based on number of subjects in each treatment group (N).", font=f_ann_sm, fill=(80,80,80))
    draw.text((L, table_bottom + 30), "Subjects counted once per SOC and once per PT, even if multiple occurrences.", font=f_ann_sm, fill=(80,80,80))

    # ── legend box ──
    lx, ly = L, H - 55
    draw.rectangle([(lx, ly), (lx+500, ly+45)], fill=(255,250,240), outline=ANN_BOX, width=1)
    draw.text((lx+6, ly+4),  "ANNOTATION LEGEND", font=f_ann, fill=ANN_TEXT)
    draw.text((lx+6, ly+18), "Red callouts = metadata used to generate R code (dataset, variables, grouping)", font=f_ann_sm, fill=ANN_TEXT)
    draw.text((lx+6, ly+32), "Upload this shell in Step 1 · Upload AdaM specs in Step 2 · Generate R code in Step 4", font=f_ann_sm, fill=ANN_TEXT)

    out = OUT_DIR / "sample_shell_annotated.png"
    img.save(str(out), dpi=(150, 150))
    print(f"Saved: {out}")


# ─────────────────────────────────────────────────────────────────
# 2. Sample AdaM specs Excel — ADAE Variable Level
# ─────────────────────────────────────────────────────────────────
def make_adam_specs():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1E3C6E")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    CELL_FONT  = Font(size=9)
    ALT_FILL   = PatternFill("solid", fgColor="F2F6FC")
    thin = Side(style="thin", color="B0B8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def hdr_row(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = wrap; c.border = border

    def data_row(ws, row, values, alt=False):
        fill = ALT_FILL if alt else PatternFill()
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = CELL_FONT; c.alignment = wrap; c.border = border
            if alt: c.fill = fill

    # ══════════════════════════════════════════════════════════════
    # Sheet 1: Variable Level
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Variable_Level"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value = "ADAE — Variable Level Metadata"
    t.font = Font(bold=True, size=12, color="1E3C6E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    ws1.merge_cells("A2:H2")
    t2 = ws1["A2"]
    t2.value = "Dataset: ADAE (Adverse Events) | Standard: CDISC ADaM v1.0 | Population: SAFFL = 'Y'"
    t2.font = Font(italic=True, size=10, color="555555")
    t2.alignment = Alignment(horizontal="center")

    cols = ["Variable", "Label", "Type", "Length", "Codelist / Values", "ADaM Core", "Key Flag", "Derivation / Notes"]
    hdr_row(ws1, 3, cols)

    variables = [
        ("STUDYID",   "Study Identifier",              "Char", 20,  "",                          "Req",    "",    "From SDTM DM.STUDYID"),
        ("USUBJID",   "Unique Subject Identifier",     "Char", 40,  "",                          "Req",    "KEY", "From SDTM DM.USUBJID — primary merge key"),
        ("SUBJID",    "Subject Identifier",            "Char", 20,  "",                          "Req",    "",    "From SDTM DM.SUBJID"),
        ("SITEID",    "Study Site Identifier",         "Char", 20,  "",                          "Perm",   "",    "From SDTM DM.SITEID"),
        ("TRTA",      "Actual Treatment",              "Char", 200, "Placebo / Drug A 50mg / Drug A 100mg", "Req", "", "Mapped from ADSL.TRT01A — use for column grouping in AE tables"),
        ("TRTP",      "Planned Treatment",             "Char", 200, "Placebo / Drug A 50mg / Drug A 100mg", "Req", "", "Mapped from ADSL.TRT01P"),
        ("AEBODSYS",  "Body System or Organ Class",    "Char", 200, "MedDRA SOC",                "Req",    "",    "System Organ Class from MedDRA coding — use for row grouping"),
        ("AEDECOD",   "Dictionary-Derived Term",       "Char", 200, "MedDRA PT",                 "Req",    "",    "Preferred Term from MedDRA coding — use for row sub-grouping under SOC"),
        ("AETERM",    "Reported Term for the AE",      "Char", 200, "",                          "Req",    "",    "Verbatim adverse event term as reported by the investigator"),
        ("AESEV",     "Severity/Intensity",            "Char", 20,  "MILD / MODERATE / SEVERE",  "Perm",   "",    "Severity grade of the adverse event"),
        ("AESER",     "Serious Event",                 "Char", 1,   "Y / N",                     "Perm",   "",    "Y = serious adverse event"),
        ("AEREL",     "Causality",                     "Char", 20,  "RELATED / NOT RELATED / POSSIBLE", "Perm", "", "Investigator assessment of relationship to study drug"),
        ("TRTEMFL",   "Treatment-Emergent Flag",       "Char", 1,   "Y",                         "Req",    "",    "Y = treatment-emergent AE. Always filter: TRTEMFL = 'Y'"),
        ("SAFFL",     "Safety Population Flag",        "Char", 1,   "Y / N",                     "Req",    "",    "Y = subject in safety population. Always filter: SAFFL = 'Y'"),
        ("ASTDT",     "Analysis Start Date",           "Num",  8,   "",                          "Perm",   "",    "Start date of adverse event (SAS date)"),
        ("AENDT",     "Analysis End Date",             "Num",  8,   "",                          "Perm",   "",    "End date of adverse event (SAS date, may be missing if ongoing)"),
    ]

    for i, row in enumerate(variables):
        data_row(ws1, i + 4, row, alt=(i % 2 == 1))

    widths = [12, 28, 6, 8, 38, 10, 8, 54]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    out = OUT_DIR / "sample_adae_spec.xlsx"
    wb.save(str(out))
    print(f"Saved: {out}")


# ─────────────────────────────────────────────────────────────────
# 3. Sample ADAE CSV — 90 subjects × 3 arms, ~3-5 AEs each
# ─────────────────────────────────────────────────────────────────
def make_adae_csv():
    random.seed(42)

    arms = ["Placebo"] * 30 + ["Drug A 50mg"] * 30 + ["Drug A 100mg"] * 30
    sites = ["001", "002", "003", "004"]

    # SOC -> list of PTs with relative weights
    ae_dict = {
        "GASTROINTESTINAL DISORDERS": [
            ("NAUSEA", 0.30), ("DIARRHEA", 0.25), ("VOMITING", 0.20),
            ("ABDOMINAL PAIN", 0.15), ("CONSTIPATION", 0.10),
        ],
        "GENERAL DISORDERS AND ADMINISTRATION SITE CONDITIONS": [
            ("FATIGUE", 0.40), ("PYREXIA", 0.30), ("OEDEMA PERIPHERAL", 0.15),
            ("ASTHENIA", 0.15),
        ],
        "NERVOUS SYSTEM DISORDERS": [
            ("HEADACHE", 0.40), ("DIZZINESS", 0.30), ("SOMNOLENCE", 0.15),
            ("PARAESTHESIA", 0.15),
        ],
        "SKIN AND SUBCUTANEOUS TISSUE DISORDERS": [
            ("RASH", 0.40), ("PRURITUS", 0.30), ("ALOPECIA", 0.15),
            ("DRY SKIN", 0.15),
        ],
        "MUSCULOSKELETAL AND CONNECTIVE TISSUE DISORDERS": [
            ("ARTHRALGIA", 0.40), ("MYALGIA", 0.30), ("BACK PAIN", 0.30),
        ],
        "INFECTIONS AND INFESTATIONS": [
            ("UPPER RESPIRATORY TRACT INFECTION", 0.35),
            ("NASOPHARYNGITIS", 0.35), ("URINARY TRACT INFECTION", 0.30),
        ],
    }

    soc_list = list(ae_dict.keys())
    severities = ["MILD", "MODERATE", "SEVERE"]
    sev_weights = [0.55, 0.35, 0.10]
    rel_values = ["NOT RELATED", "POSSIBLE", "RELATED"]
    rel_weights = [0.50, 0.30, 0.20]

    # Drug arms get more AEs on average
    ae_count_range = {
        "Placebo":      (1, 4),
        "Drug A 50mg":  (2, 5),
        "Drug A 100mg": (2, 6),
    }

    header = [
        "STUDYID", "DOMAIN", "USUBJID", "SUBJID", "SITEID",
        "TRTA", "TRTP", "AEBODSYS", "AEDECOD", "AETERM",
        "AESEV", "AESER", "AEREL", "TRTEMFL", "SAFFL",
        "ASTDT", "AENDT",
    ]

    rows = []
    base_date = datetime(2023, 10, 1)

    for subj_idx, arm in enumerate(arms, 1):
        subjid = f"{subj_idx:04d}"
        site = random.choice(sites)
        usubjid = f"MYSTUDY01-{site}-{subjid}"

        n_aes = random.randint(*ae_count_range[arm])
        # Pick random SOCs for this subject's AEs
        chosen_socs = random.choices(soc_list, k=n_aes)

        for ae_i, soc in enumerate(chosen_socs):
            pts, pt_wts = zip(*ae_dict[soc])
            pt = random.choices(pts, weights=pt_wts, k=1)[0]

            sev = random.choices(severities, weights=sev_weights, k=1)[0]
            ser = "Y" if sev == "SEVERE" and random.random() < 0.5 else "N"
            rel = random.choices(rel_values, weights=rel_weights, k=1)[0]

            start = base_date + timedelta(days=random.randint(1, 90))
            end = start + timedelta(days=random.randint(1, 30))
            # ~10% ongoing (missing end date)
            end_str = end.strftime("%Y-%m-%d") if random.random() > 0.10 else ""

            rows.append([
                "MYSTUDY01", "ADAE", usubjid, subjid, site,
                arm, arm, soc, pt, pt,
                sev, ser, rel, "Y", "Y",
                start.strftime("%Y-%m-%d"), end_str,
            ])

    out = OUT_DIR / "sample_adae.csv"
    with open(out, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print(f"Saved: {out}  ({len(rows)} records, {len(arms)} subjects)")


# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    make_annotated_shell()
    make_adam_specs()
    make_adae_csv()
    print("Sample files created successfully.")
